import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os, shutil
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

# ---------------- CONFIG ----------------
DATA_DIR = "MoiData"
BACKUP_DIR = "backup"
LOCK_FILE = "function.lock"

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# ---------------- LANGUAGE ----------------
LANG = "TA"

TEXT = {
    "TA": {
        "title": "மொய் கணக்கு",
        "name": "பெயர்",
        "guardian": "தந்தை / துணை பெயர்",
        "address": "முகவரி",
        "amount": "தொகை",
        "function": "நிகழ்ச்சி பெயர்",
        "save": "சேமிக்க",
        "total": "மொத்த தொகை",
        "edit": "கடைசி 5 பதிவுகளை திருத்து",
        "toggle": "English",
        "required": "பெயர், முகவரி, தொகை அவசியம்"
    },
    "EN": {
        "title": "MoiKanakku",
        "name": "Name",
        "guardian": "Father / Spouse",
        "address": "Address",
        "amount": "Amount",
        "function": "Function Name",
        "save": "Save",
        "total": "Total",
        "edit": "Edit Last 5 Records",
        "toggle": "தமிழ்",
        "required": "Name, Address and Amount are required"
    }
}

def t(key):
    return TEXT[LANG][key]

# ---------------- FUNCTION FILE ----------------
def get_excel_file(function):
    date = datetime.now().strftime("%Y-%m-%d")
    safe = function.replace(" ", "_")
    return os.path.join(DATA_DIR, f"{date}_{safe}.xlsx")

def create_excel(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Guardian", "Address", "Amount", "Date"])
        wb.save(path)

# ---------------- AUTOSUGGEST ----------------
class AutoEntry(ctk.CTkEntry):
    def __init__(self, master, values, **kwargs):
        super().__init__(master, **kwargs)
        self.values = values
        self.lb = None
        self.bind("<KeyRelease>", self.show)

    def show(self, _=None):
        if self.lb:
            self.lb.destroy()
        text = self.get().lower()
        if not text:
            return
        matches = [v for v in self.values if v.lower().startswith(text)]
        if not matches:
            return
        self.lb = tk.Listbox(self.master, height=min(5, len(matches)))
        self.lb.place(x=self.winfo_x(), y=self.winfo_y() + 30)
        for m in matches:
            self.lb.insert(tk.END, m)
        self.lb.bind("<<ListboxSelect>>", self.select)

    def select(self, _):
        self.delete(0, tk.END)
        self.insert(0, self.lb.get(self.lb.curselection()))
        self.lb.destroy()

# ---------------- LOAD SUGGESTIONS ----------------
def load_suggestions():
    names, guardians, addresses = set(), set(), set()
    for file in os.listdir(DATA_DIR):
        if file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(DATA_DIR, file))
            ws = wb.active
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0]: names.add(r[0])
                if r[1]: guardians.add(r[1])
                if r[2]: addresses.add(r[2])
    return list(names), list(guardians), list(addresses)

name_s, guardian_s, address_s = load_suggestions()

# ---------------- LOAD LOCK ----------------
locked_function = ""
if os.path.exists(LOCK_FILE):
    locked_function = open(LOCK_FILE).read().strip()

# ---------------- APP ----------------
ctk.set_appearance_mode("light")
app = ctk.CTk()
app.geometry("520x600")

# ---------------- FUNCTIONS ----------------
def refresh_ui():
    app.title(t("title"))
    title.configure(text=t("title"))
    name_lbl.configure(text=t("name"))
    guardian_lbl.configure(text=t("guardian"))
    address_lbl.configure(text=t("address"))
    amount_lbl.configure(text=t("amount"))
    function_lbl.configure(text=t("function"))
    save_btn.configure(text=t("save"))
    total_btn.configure(text=t("total"))
    edit_btn.configure(text=t("edit"))
    toggle_btn.configure(text=t("toggle"))

def toggle_language():
    global LANG
    LANG = "EN" if LANG == "TA" else "TA"
    refresh_ui()

def save():
    global locked_function
    name = name_e.get()
    guardian = guardian_e.get()
    address = address_e.get()
    amount = amount_e.get()
    function = function_e.get()

    if not name or not address or not amount:
        messagebox.showwarning("Error", t("required"))
        return

    if not locked_function:
        locked_function = function
        open(LOCK_FILE, "w").write(function)
        function_e.configure(state="disabled")

    file = get_excel_file(locked_function)
    create_excel(file)

    wb = load_workbook(file)
    ws = wb.active
    ws.append([name, guardian, address, int(amount), datetime.now().strftime("%d-%m-%Y %H:%M")])
    wb.save(file)

    shutil.copy(file, os.path.join(BACKUP_DIR, os.path.basename(file)))

    name_e.delete(0, tk.END)
    guardian_e.delete(0, tk.END)
    address_e.delete(0, tk.END)
    amount_e.delete(0, tk.END)

def show_total():
    file = get_excel_file(locked_function)
    wb = load_workbook(file)
    ws = wb.active
    total = sum([r[3] for r in ws.iter_rows(min_row=2, values_only=True)])
    messagebox.showinfo(t("total"), f"₹ {total}")

def edit_last5():
    file = get_excel_file(locked_function)
    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2))
    last = rows[-5:]

    win = tk.Toplevel()
    win.title("Edit Last 5")

    for i, r in enumerate(last):
        tk.Entry(win, width=20).insert(0, r[0].value)

# ---------------- UI ----------------
title = ctk.CTkLabel(app, font=("Arial", 22, "bold"))
title.pack(pady=10)

toggle_btn = ctk.CTkButton(app, width=90, command=toggle_language)
toggle_btn.pack(anchor="e", padx=20)

form = ctk.CTkFrame(app)
form.pack(pady=10, padx=20, fill="x")

def row(label, widget):
    label.pack(anchor="w")
    widget.pack(fill="x", pady=5)

name_lbl = ctk.CTkLabel(form)
name_e = AutoEntry(form, name_s)
row(name_lbl, name_e)

guardian_lbl = ctk.CTkLabel(form)
guardian_e = AutoEntry(form, guardian_s)
row(guardian_lbl, guardian_e)

address_lbl = ctk.CTkLabel(form)
address_e = AutoEntry(form, address_s)
row(address_lbl, address_e)

amount_lbl = ctk.CTkLabel(form)
amount_e = ctk.CTkEntry(form)
row(amount_lbl, amount_e)

function_lbl = ctk.CTkLabel(form)
function_e = ctk.CTkEntry(form)
row(function_lbl, function_e)

if locked_function:
    function_e.insert(0, locked_function)
    function_e.configure(state="disabled")

save_btn = ctk.CTkButton(app, command=save)
save_btn.pack(pady=10)

total_btn = ctk.CTkButton(app, command=show_total)
total_btn.pack(pady=5)

edit_btn = ctk.CTkButton(app, command=edit_last5)
edit_btn.pack(pady=5)

refresh_ui()
app.mainloop()
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os, shutil
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

# ---------- CONFIG ----------
DATA_DIR = "MoiData"
BACKUP_DIR = "backup"
LOCK_FILE = "function.lock"

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# ---------- LANGUAGE ----------
LANG = "TA"

TEXT = {
    "TA": {
        "title": "மொய் கணக்கு",
        "name": "பெயர்",
        "guardian": "தந்தை / துணை",
        "address": "முகவரி",
        "amount": "தொகை",
        "function": "நிகழ்ச்சி பெயர்",
        "save": "சேமிக்க",
        "total": "மொத்த தொகை",
        "edit": "கடைசி 5 பதிவுகளை திருத்து",
        "new": "புதிய நிகழ்ச்சி",
        "toggle": "English",
        "required": "பெயர், முகவரி மற்றும் தொகை அவசியம்"
    },
    "EN": {
        "title": "MoiKanakku",
        "name": "Name",
        "guardian": "Father / Spouse",
        "address": "Address",
        "amount": "Amount",
        "function": "Function Name",
        "save": "Save",
        "total": "Total",
        "edit": "Edit Last 5",
        "new": "New Function",
        "toggle": "தமிழ்",
        "required": "Name, Address and Amount are required"
    }
}

def t(key):
    return TEXT[LANG][key]

# ---------- FILE ----------
def get_excel_file(function):
    date = datetime.now().strftime("%Y-%m-%d")
    safe = function.replace(" ", "_")
    return os.path.join(DATA_DIR, f"{date}_{safe}.xlsx")

def create_excel(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Guardian", "Address", "Amount", "Date"])
        wb.save(path)

# ---------- AUTOSUGGEST ----------
class AutoEntry(ctk.CTkEntry):
    def __init__(self, master, values, **kw):
        super().__init__(master, **kw)
        self.values = values
        self.lb = None
        self.bind("<KeyRelease>", self.show)

    def show(self, _=None):
        if self.lb:
            self.lb.destroy()
        t = self.get().lower()
        if not t:
            return
        m = [v for v in self.values if v.lower().startswith(t)]
        if not m:
            return
        self.lb = tk.Listbox(self.master, height=min(5, len(m)))
        self.lb.place(x=self.winfo_x(), y=self.winfo_y() + 30)
        for i in m:
            self.lb.insert(tk.END, i)
        self.lb.bind("<<ListboxSelect>>", self.select)

    def select(self, _):
        self.delete(0, tk.END)
        self.insert(0, self.lb.get(self.lb.curselection()))
        self.lb.destroy()

# ---------- LOAD SUGGESTIONS ----------
def load_suggestions():
    names, guardians, addresses = set(), set(), set()
    for file in os.listdir(DATA_DIR):
        if file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(DATA_DIR, file))
            ws = wb.active
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0]: names.add(r[0])
                if r[1]: guardians.add(r[1])
                if r[2]: addresses.add(r[2])
    return list(names), list(guardians), list(addresses)

name_s, guardian_s, address_s = load_suggestions()

# ---------- LOCK ----------
locked_function = ""
if os.path.exists(LOCK_FILE):
    locked_function = open(LOCK_FILE).read().strip()

# ---------- APP ----------
ctk.set_appearance_mode("light")
app = ctk.CTk()
app.geometry("520x640")

# ---------- FUNCTIONS ----------
def refresh_ui():
    app.title(t("title"))
    title.configure(text=t("title"))
    name_lbl.configure(text=t("name"))
    guardian_lbl.configure(text=t("guardian"))
    address_lbl.configure(text=t("address"))
    amount_lbl.configure(text=t("amount"))
    function_lbl.configure(text=t("function"))
    save_btn.configure(text=t("save"))
    total_btn.configure(text=t("total"))
    edit_btn.configure(text=t("edit"))
    new_btn.configure(text=t("new"))
    toggle_btn.configure(text=t("toggle"))

def toggle_lang():
    global LANG
    LANG = "EN" if LANG == "TA" else "TA"
    refresh_ui()

def save():
    global locked_function
    name = name_e.get()
    guardian = guardian_e.get()
    address = address_e.get()
    amount = amount_e.get()
    function = function_e.get()

    if not name or not address or not amount:
        messagebox.showwarning("Error", t("required"))
        return

    if not locked_function:
        locked_function = function
        open(LOCK_FILE, "w").write(function)
        function_e.configure(state="disabled")

    file = get_excel_file(locked_function)
    create_excel(file)

    wb = load_workbook(file)
    ws = wb.active
    ws.append([name, guardian, address, int(amount), datetime.now().strftime("%d-%m-%Y %H:%M")])
    wb.save(file)

    shutil.copy(file, os.path.join(BACKUP_DIR, os.path.basename(file)))

    name_e.delete(0, tk.END)
    guardian_e.delete(0, tk.END)
    address_e.delete(0, tk.END)
    amount_e.delete(0, tk.END)

def new_function():
    global locked_function
    if messagebox.askyesno("Confirm", "Start new function?"):
        locked_function = ""
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
        function_e.configure(state="normal")
        function_e.delete(0, tk.END)

def show_total():
    file = get_excel_file(locked_function)
    wb = load_workbook(file)
    ws = wb.active
    total = sum([r[3] for r in ws.iter_rows(min_row=2, values_only=True)])
    messagebox.showinfo(t("total"), f"₹ {total}")

def edit_last5():
    file = get_excel_file(locked_function)
    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2))
    last = rows[-5:]

    win = tk.Toplevel()
    win.title("Edit Last 5")

    entries = []
    for i, r in enumerate(last):
        e1 = tk.Entry(win); e1.insert(0, r[0].value); e1.grid(row=i, column=0)
        e2 = tk.Entry(win); e2.insert(0, r[2].value); e2.grid(row=i, column=1)
        e3 = tk.Entry(win); e3.insert(0, r[3].value); e3.grid(row=i, column=2)
        entries.append((r, e1, e2, e3))

    def save_edits():
        for r, e1, e2, e3 in entries:
            r[0].value = e1.get()
            r[2].value = e2.get()
            r[3].value = int(e3.get())
        wb.save(file)
        messagebox.showinfo("Saved", "Updated successfully")
        win.destroy()

    tk.Button(win, text="Save", command=save_edits).grid(row=len(entries), column=1)

# ---------- UI ----------
title = ctk.CTkLabel(app, font=("Arial", 22, "bold"))
title.pack(pady=10)

toggle_btn = ctk.CTkButton(app, width=90, command=toggle_lang)
toggle_btn.pack(anchor="e", padx=20)

form = ctk.CTkFrame(app)
form.pack(padx=20, pady=10, fill="x")

def field(lbl, entry):
    lbl.pack(anchor="w")
    entry.pack(fill="x", pady=5)

name_lbl = ctk.CTkLabel(form)
name_e = AutoEntry(form, name_s)
field(name_lbl, name_e)

guardian_lbl = ctk.CTkLabel(form)
guardian_e = AutoEntry(form, guardian_s)
field(guardian_lbl, guardian_e)

address_lbl = ctk.CTkLabel(form)
address_e = AutoEntry(form, address_s)
field(address_lbl, address_e)

amount_lbl = ctk.CTkLabel(form)
amount_e = ctk.CTkEntry(form)
field(amount_lbl, amount_e)

function_lbl = ctk.CTkLabel(form)
function_e = ctk.CTkEntry(form)
field(function_lbl, function_e)

if locked_function:
    function_e.insert(0, locked_function)
    function_e.configure(state="disabled")

save_btn = ctk.CTkButton(app, command=save)
save_btn.pack(pady=10)

total_btn = ctk.CTkButton(app, command=show_total)
total_btn.pack()

edit_btn = ctk.CTkButton(app, command=edit_last5)
edit_btn.pack(pady=5)

new_btn = ctk.CTkButton(app, command=new_function)
new_btn.pack(pady=5)

refresh_ui()
app.mainloop()
